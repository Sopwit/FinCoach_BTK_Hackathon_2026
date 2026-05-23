import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import crud, schemas
from app.database import get_db

router = APIRouter(
    prefix="/transactions",
    tags=["Transactions"]
)


def parse_transaction_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        raise ValueError("date empty")

    for date_format in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue

    try:
        return datetime.fromisoformat(text).date()
    except (ValueError, TypeError):
        raise ValueError(f"Tarih formatı tanınamadı: {text}")


def normalize_upload_row(row: dict) -> dict:
    return {
        str(key).strip().lower(): value
        for key, value in row.items()
        if key is not None
    }


def parse_csv_transactions(content: bytes) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1254")

    reader = csv.DictReader(StringIO(text))
    return [normalize_upload_row(row) for row in reader]


def parse_xlsx_transactions(content: bytes) -> list[dict]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)

    if not headers:
        return []

    normalized_headers = [str(header).strip().lower() if header is not None else "" for header in headers]
    return [
        normalize_upload_row(dict(zip(normalized_headers, row)))
        for row in rows
    ]


@router.post("/manual", response_model=schemas.TransactionResponse)
def create_manual_transaction(
    transaction: schemas.TransactionCreate,
    db: Session = Depends(get_db)
):
    user = crud.get_user(db=db, user_id=transaction.user_id)

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    return crud.create_transaction(db=db, transaction=transaction)

@router.get("/", response_model=List[schemas.TransactionResponse])
def list_transactions(
    user_id: Optional[int] = Query(default=None),
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    tx_type: Optional[str] = Query(default=None, alias="type"),
    category: Optional[str] = Query(default=None),
    source: Optional[str] = Query(default=None),
    search: Optional[str] = Query(default=None),
    db: Session = Depends(get_db)
):
    return crud.get_transactions(
        db=db,
        user_id=user_id,
        year=year,
        month=month,
        tx_type=tx_type,
        category=category,
        source=source,
        search=search
    )
@router.post("/upload", response_model=schemas.BulkTransactionResponse)
async def upload_transactions_file(
    user_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    user = crud.get_user(db=db, user_id=user_id)

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    filename = file.filename.lower()
    content = await file.read()

    try:
        if filename.endswith(".csv"):
            rows = parse_csv_transactions(content)
        elif filename.endswith(".xlsx"):
            rows = parse_xlsx_transactions(content)
        else:
            raise HTTPException(
                status_code=400,
                detail="Only CSV or Excel files are supported"
            )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"File could not be read: {str(e)}"
        )

    required_columns = {"date", "description", "amount", "type"}

    if not rows or not required_columns.issubset(set(rows[0].keys())):
        raise HTTPException(
            status_code=400,
            detail="File must contain these columns: date, description, amount, type"
        )

    transactions_to_create = []

    for row in rows:
        try:
            tx_date = parse_transaction_date(row["date"])
            description = str(row["description"])
            amount = float(row["amount"])
            tx_type = str(row["type"]).strip().lower()

            if tx_type not in ["income", "expense"]:
                continue

            transaction = schemas.TransactionCreate(
                user_id=user_id,
                date=tx_date,
                description=description,
                amount=amount,
                type=tx_type,
                source="csv"
            )

            transactions_to_create.append(transaction)

        except Exception:
            continue

    result = crud.create_many_transactions(
    db=db,
    transactions=transactions_to_create,
    skip_duplicates=True
)

    created_transactions = result["created_transactions"]
    skipped_count = result["skipped_count"]

    return {
    "inserted_count": len(created_transactions),
    "skipped_count": skipped_count,
    "transactions": created_transactions
    }


@router.delete("/")
def delete_transactions(
    user_id: int = Query(...),
    year: Optional[int] = Query(default=None),
    month: Optional[int] = Query(default=None),
    db: Session = Depends(get_db)
):
    user = crud.get_user(db=db, user_id=user_id)

    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    deleted_count = crud.delete_transactions(
        db=db,
        user_id=user_id,
        year=year,
        month=month
    )

    return {
        "message": "Transactions deleted successfully",
        "deleted_count": deleted_count
    }


@router.put("/{transaction_id}", response_model=schemas.TransactionResponse)
def update_transaction(
    transaction_id: int,
    transaction_update: schemas.TransactionUpdate,
    db: Session = Depends(get_db)
):
    try:
        updated_transaction = crud.update_transaction(
            db=db,
            transaction_id=transaction_id,
            transaction_update=transaction_update
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not updated_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    return updated_transaction


@router.delete("/{transaction_id}")
def delete_transaction(
    transaction_id: int,
    db: Session = Depends(get_db)
):
    deleted_transaction = crud.delete_transaction(
        db=db,
        transaction_id=transaction_id
    )

    if not deleted_transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")

    return {
        "message": "Transaction deleted successfully",
        "deleted_transaction_id": transaction_id
    }
